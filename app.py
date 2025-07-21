import streamlit as st
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options
from webdriver_manager.microsoft import EdgeChromiumDriverManager

import pandas as pd
import os
import time
import glob
from datetime import timedelta
import platform
import tempfile
import shutil

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ======================================================================================
# PATHS
# ======================================================================================
DOWNLOAD_DIR = os.path.join(os.getcwd(), "downloads_temp")
FINAL_DIR = os.path.join(os.getcwd(), "consolidated_reports")
EXPORT_FILE = os.path.join(os.getcwd(), "clients_list.csv")

os.makedirs(DOWNLOAD_DIR, exist_ok=True)
os.makedirs(FINAL_DIR, exist_ok=True)


# ======================================================================================
# UTILITIES
# ======================================================================================
def parse_time(t):
    """Parse mixed time formats to timedelta."""
    try:
        t = str(t).strip().lower()
        if not t:
            return timedelta(0)

        h, m = 0, 0
        if ":" in t:
            h, m = map(int, t.split(":"))
        elif "." in t:  # e.g., 7.5 hrs
            parts = t.split(".")
            h = int(parts[0])
            m = int(float("0." + parts[1]) * 60)
        else:
            h = int(float(t))
        return timedelta(hours=h, minutes=m)
    except Exception:
        return timedelta(0)


def format_td(td):
    total_minutes = int(td.total_seconds() // 60)
    return f"{total_minutes // 60}h {total_minutes % 60}m"


def style_excel(file_path: str):
    """Apply simple styling to a 1-sheet summary workbook."""
    wb = load_workbook(file_path)
    ws = wb.active
    ws.title = "Summary"

    ws.freeze_panes = ws["A2"]
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(border_style="thin", color="000000")

    # Header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Body rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.row % 2 == 0:
                cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Autosize
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    wb.save(file_path)


def safe_rename(src_path: str, dest_path: str) -> str:
    """Rename without clobbering existing file. Returns final path."""
    base, ext = os.path.splitext(dest_path)
    counter = 1
    final_path = dest_path
    while os.path.exists(final_path):
        final_path = f"{base}_{counter}{ext}"
        counter += 1
    os.rename(src_path, final_path)
    return final_path


# ======================================================================================
# SELENIUM DRIVER FACTORY
# ======================================================================================
def create_edge_driver(headless: bool = True):
    """Create and return a configured Edge WebDriver."""
    options = Options()
    options.use_chromium = True
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")

    if headless:
        # Use the most compatible headless arg per platform
        if platform.system() == "Windows":
            options.add_argument("--headless=new")
        else:
            options.add_argument("--headless=new")
        # Helpful for remote / container runs
        options.add_argument("--window-size=1920,1080")
        options.add_argument("--disable-gpu")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")

    # Download prefs
    options.add_experimental_option(
        "prefs",
        {
            "download.default_directory": DOWNLOAD_DIR,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True,
            "profile.default_content_settings.popups": 0,
        },
    )

    driver = webdriver.Edge(
        service=EdgeService(EdgeChromiumDriverManager().install()),
        options=options,
    )

    # Enable downloads in headless mode via CDP
    try:
        driver.execute_cdp_cmd(
            "Page.setDownloadBehavior",
            {"behavior": "allow", "downloadPath": DOWNLOAD_DIR},
        )
    except Exception as e:
        # Non-fatal; downloads may still work
        print(f"[WARN] Could not enable CDP download override: {e}")

    return driver


# ======================================================================================
# WORKFLOW: DOWNLOAD TIMESHEETS
# ======================================================================================
def download_timesheets():
    st.title("📥 Download Timesheets")
    if st.button("⬅️ Back to Home"):
        st.session_state.page = "home"
        st.rerun()

    username = st.text_input("👤 Username")
    password = st.text_input("🔑 Password", type="password")
    from_date = st.text_input("📅 FROM Date (DD/MM/YYYY)")
    to_date = st.text_input("📅 TO Date (DD/MM/YYYY)")
    headless_mode = st.toggle("Headless Mode", value=True)

    st.caption("Downloads will be saved in ./downloads_temp")

    start_button = st.button("🚀 Start Automation")

    if not (start_button and username and password and from_date and to_date):
        st.info("⬆️ Fill all inputs and press Start Automation")
        return

    # ------------------------------------------------------------------
    # Load clients CSV
    # ------------------------------------------------------------------
    try:
        df_clients = pd.read_csv(EXPORT_FILE)
    except FileNotFoundError:
        st.error("❌ clients_list.csv not found in current directory.")
        return

    if "Client" not in df_clients.columns:
        st.error("❌ 'Client' column not found in clients_list.csv.")
        return

    df_clients["Status"] = ""

    clients = df_clients["Client"].dropna().astype(str).tolist()
    n_clients = len(clients)

    progress_text = st.empty()
    progress_bar = st.progress(0.0)
    log_box = st.empty()

    # Optional: Clean download dir to avoid confusion with stale files
    # for f in glob.glob(os.path.join(DOWNLOAD_DIR, "*")):
    #     try:
    #         os.remove(f)
    #     except Exception:
    #         pass

    # ------------------------------------------------------------------
    # Process each client independently (new browser per client)
    # (Safer against session state bleed; change to single-driver loop if desired.)
    # ------------------------------------------------------------------
    for idx, client in enumerate(clients, start=1):
        progress_text.text(f"🔄 Processing [{idx}/{n_clients}]: {client}")
        progress_bar.progress(idx / n_clients)
        log_box.write(f"---\n**Client:** {client}")

        driver = None  # <-- guarantee name exists for finally block
        try:
            driver = create_edge_driver(headless=headless_mode)
            wait = WebDriverWait(driver, 30)

            # ---------------------------------
            # Login
            # ---------------------------------
            driver.get("https://timesheet.outsourcinghubindia.com")
            wait.until(
                EC.presence_of_element_located(
                    (By.XPATH, "//*[@id='l-login']/form/div/div[1]/div/input")
                )
            ).send_keys(username)
            wait.until(
                EC.presence_of_element_located(
                    (By.XPATH, "//*[@id='l-login']/form/div/div[2]/div/input")
                )
            ).send_keys(password)
            wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//*[@id='l-login']/form/div/button")
                )
            ).click()

            # ---------------------------------
            # Navigate to Reports/Export page
            # ---------------------------------
            wait.until(
                EC.element_to_be_clickable((By.XPATH, "/html/body/header/div[2]/div/div/div/ul/li[5]/a"))
            ).click()
            wait.until(
                EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div/div/div[1]/a"))
            ).click()

            # ---------------------------------
            # Select client in dropdown
            # ---------------------------------
            time.sleep(1)
            process_dropdown = wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, "/html/body/div[2]/div/div/div/div/form/div/div[1]/div/div/div/button")
                )
            )
            process_dropdown.click()

            search_input = wait.until(
                EC.presence_of_element_located(
                    (By.XPATH, "/html/body/div[2]/div/div/div/div/form/div/div[1]/div/div/div/div/div/input")
                )
            )
            search_input.clear()
            search_input.send_keys(client)
            time.sleep(1)

            # Dropdown list items
            client_elements = wait.until(
                EC.presence_of_all_elements_located(
                    (By.CSS_SELECTOR, "ul.dropdown-menu.inner span.text")
                )
            )
            found = False
            for element in client_elements:
                if element.text.strip().lower() == client.lower():
                    element.click()
                    found = True
                    break

            if not found:
                st.warning(f"⚠️ {client}: Not found in dropdown. Skipping.")
                df_clients.loc[df_clients["Client"] == client, "Status"] = "Skipped - Not Found"
                continue

            # ---------------------------------
            # Dates
            # ---------------------------------
            from_box = wait.until(
                EC.presence_of_element_located(
                    (By.XPATH, "/html/body/div[2]/div/div/div/div/form/div/div[3]/div/div[1]/div/input")
                )
            )
            to_box = wait.until(
                EC.presence_of_element_located(
                    (By.XPATH, "/html/body/div[2]/div/div/div/div/form/div/div[3]/div/div[2]/div/input")
                )
            )
            from_box.clear()
            to_box.clear()
            from_box.send_keys(from_date)
            to_box.send_keys(to_date)

            wait.until(EC.element_to_be_clickable((By.XPATH, "//button[text()='Search']"))).click()

            # ---------------------------------
            # Check if data rows exist
            # ---------------------------------
            try:
                WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, "//table//tbody/tr"))
                )
            except Exception:
                st.warning(f"⚠️ No data for {client}. Skipping.")
                df_clients.loc[df_clients["Client"] == client, "Status"] = "Skipped - No Data"
                continue

            # ---------------------------------
            # Detect new download
            # ---------------------------------
            existing_files = set(glob.glob(os.path.join(DOWNLOAD_DIR, "*.xlsx")))

            # Click export
            wait.until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(text(),'Export XLS')]"))
            ).click()
            log_box.write(f"Clicked Export XLS for **{client}**; waiting for file...")

            # Wait for up to 45s (longer to be safer in cloud environments)
            download_complete = False
            latest_path = None
            for _ in range(45):
                time.sleep(1)
                current_files = set(glob.glob(os.path.join(DOWNLOAD_DIR, "*.xlsx")))
                new_files = current_files - existing_files
                if new_files:
                    # Pick newest
                    latest_path = max(new_files, key=os.path.getctime)
                    download_complete = True
                    break

            if download_complete and latest_path:
                new_name = os.path.join(DOWNLOAD_DIR, f"{client}.xlsx")
                final_path = safe_rename(latest_path, new_name)
                st.success(f"✅ Downloaded: {os.path.basename(final_path)}")
                df_clients.loc[df_clients["Client"] == client, "Status"] = "Downloaded"
            else:
                st.error(f"❌ Download failed for {client}")
                df_clients.loc[df_clients["Client"] == client, "Status"] = "Failed Download"

        except Exception as e:
            st.error(f"❌ Error processing '{client}': {e}")
            df_clients.loc[df_clients["Client"] == client, "Status"] = "Error"

        finally:
            # SAFELY quit
            if driver is not None:
                try:
                    driver.quit()
                except Exception as e:
                    print(f"[WARN] driver.quit() failed for {client}: {e}")
            time.sleep(0.25)

    # ------------------------------------------------------------------
    # Wrap up
    # ------------------------------------------------------------------
    progress_text.text("🎉 Completed all clients!")
    progress_bar.progress(1.0)

    status_file = os.path.join(FINAL_DIR, "clients_with_status.csv")
    df_clients.to_csv(status_file, index=False)
    st.success(f"📄 Status file saved to: {status_file}")


# ======================================================================================
# WORKFLOW: CONSOLIDATE EXCELS
# ======================================================================================
def consolidate_excels():
    st.title("📊 Consolidate Excel Files")
    if st.button("⬅️ Back to Home"):
        st.session_state.page = "home"
        st.rerun()

    uploaded_files = st.file_uploader(
        "Upload Excel Files",
        type="xlsx",
        accept_multiple_files=True,
        help="Select one or more downloaded timesheet XLSX files.",
    )

    if not uploaded_files:
        st.info("📂 Upload Excel files to start.")
        return

    all_data = []
    for file in uploaded_files:
        try:
            # Timesheet exports usually have headers after 3 rows; adjust if needed.
            df = pd.read_excel(file, header=3)
            df.columns = [str(c).strip() for c in df.columns]

            required = {"Employee", "Process Name", "Work Type"}
            if not required.issubset(df.columns):
                st.warning(f"⚠️ Skipping {file.name}: Missing required columns ({required}).")
                continue

            possible_time_cols = [c for c in df.columns if "time" in c.lower()]
            time_col = next(
                (c for c in possible_time_cols if "hrs" in c.lower() or "hour" in c.lower()),
                None,
            )
            if not time_col:
                st.warning(f"⚠️ Skipping {file.name}: No time column found.")
                continue

            df[time_col] = df[time_col].apply(parse_time)
            df.rename(columns={time_col: "Time Worked"}, inplace=True)
            all_data.append(df)

        except Exception as e:
            st.error(f"❌ Error reading {file.name}: {e}")

    if not all_data:
        st.error("❌ No valid files found.")
        return

    combined_df = pd.concat(all_data, ignore_index=True)
    summary_df = (
        combined_df.groupby(["Employee", "Process Name", "Work Type"])["Time Worked"]
        .sum()
        .reset_index()
    )
    summary_df["Total Time Worked"] = summary_df["Time Worked"].apply(format_td)
    summary_df.drop(columns=["Time Worked"], inplace=True)

    st.dataframe(summary_df, use_container_width=True)

    tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    with pd.ExcelWriter(tmp_file.name, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False)
    style_excel(tmp_file.name)

    with open(tmp_file.name, "rb") as f:
        st.download_button("⬇️ Download Excel", f, file_name="Consolidated_Summary.xlsx")


# ======================================================================================
# MAIN APP ROUTER
# ======================================================================================
if "page" not in st.session_state:
    st.session_state.page = "home"

if st.session_state.page == "home":
    st.title("📑 Timesheet Automation & Consolidation")

    col1, col2 = st.columns(2)
    with col1:
        if st.button("📥 Download Timesheets"):
            st.session_state.page = "download"
            st.rerun()
    with col2:
        if st.button("📊 Consolidate Excel Files"):
            st.session_state.page = "consolidate"
            st.rerun()

elif st.session_state.page == "download":
    download_timesheets()

elif st.session_state.page == "consolidate":
    consolidate_excels()
